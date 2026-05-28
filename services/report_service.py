"""
ReportService — render an AnalysisJob into PDF or Excel.

Uses only free libraries: ReportLab for PDF, openpyxl for XLSX.
matplotlib + pandapower.plotting are used for single-line diagram
images embedded in PDFs.

The generated file is written under app.config['REPORT_FOLDER']
and the path is stored on the Report row.
"""
from __future__ import annotations

import io
import os
from datetime import datetime, timezone
from typing import Any

from flask import current_app

from extension import db
from Models import AnalysisJob, Report


class ReportService:

    # =================================================================
    #  Entry point — picks PDF or XLSX
    # =================================================================
    @classmethod
    def generate(cls, report_id: int, include_diagrams: bool = True) -> Report:
        report = db.session.get(Report, report_id)
        if report is None:
            raise ValueError(f"Report {report_id} not found")
        job = db.session.get(AnalysisJob, report.job_id)
        if job is None:
            raise ValueError(f"AnalysisJob {report.job_id} not found")

        base_dir = current_app.config.get("REPORT_FOLDER", "reports")
        os.makedirs(base_dir, exist_ok=True)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"report_{report.id}_{job.analysis_type.value}_{ts}.{report.format}"
        full_path = os.path.join(base_dir, filename)

        if report.format == "pdf":
            cls._render_pdf(job, full_path, include_diagrams=include_diagrams)
        elif report.format == "xlsx":
            cls._render_xlsx(job, full_path)
        else:
            raise ValueError(f"Unsupported format: {report.format}")

        report.file_path       = filename
        report.file_size_bytes = os.path.getsize(full_path)
        db.session.commit()
        return report

    # =================================================================
    #  PDF (ReportLab)
    # =================================================================
    @classmethod
    def _render_pdf(cls, job: AnalysisJob, out_path: str,
                    include_diagrams: bool = True) -> None:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image,
        )

        doc = SimpleDocTemplate(
            out_path, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm,
            topMargin=2 * cm, bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title2", parent=styles["Title"], fontSize=20, spaceAfter=20,
            textColor=colors.HexColor("#1f4e79"),
        )
        h2 = ParagraphStyle(
            "H2", parent=styles["Heading2"], fontSize=14,
            textColor=colors.HexColor("#1f4e79"), spaceBefore=14, spaceAfter=8,
        )
        normal = styles["BodyText"]

        elements = []
        # Title
        elements.append(Paragraph(
            f"{job.analysis_type.value.replace('_', ' ').title()} Report",
            title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            normal,
        ))
        elements.append(Spacer(1, 12))

        # Job metadata
        elements.append(Paragraph("Analysis Details", h2))
        meta = [
            ["Job ID",        str(job.id)],
            ["Analysis Type", job.analysis_type.value],
            ["Status",        job.status.value],
            ["Network ID",    str(job.network_id)],
            ["Converged",     str(job.converged) if job.converged is not None else "—"],
            ["Duration",
                f"{job.duration_sec:.2f} s" if job.duration_sec else "—"],
            ["Started",
                job.started_at.isoformat() if job.started_at else "—"],
            ["Completed",
                job.completed_at.isoformat() if job.completed_at else "—"],
        ]
        elements.append(cls._table(meta, [5 * cm, 11 * cm]))
        elements.append(Spacer(1, 12))

        # Summary
        results = job.results or {}
        summary = results.get("summary") or {}
        if summary:
            elements.append(Paragraph("Summary", h2))
            rows = [[k.replace("_", " ").title(),
                     f"{v:.4f}" if isinstance(v, float) else str(v)]
                    for k, v in summary.items()]
            elements.append(cls._table(rows, [9 * cm, 7 * cm]))
            elements.append(Spacer(1, 12))

        # Violations
        if job.violations:
            elements.append(Paragraph(
                f"Violations ({len(job.violations)})", h2))
            v_rows = [["Element", "Type", "Severity", "Value", "Limit", "Unit"]]
            for v in job.violations[:50]:
                v_rows.append([
                    f"{v.element_type.value} #{v.element_pp_index}",
                    v.violation_type, v.severity.value,
                    f"{v.value:.3f}" if v.value is not None else "—",
                    f"{v.limit:.3f}" if v.limit is not None else "—",
                    v.unit or "",
                ])
            elements.append(cls._table(v_rows, header_row=True))
            if len(job.violations) > 50:
                elements.append(Paragraph(
                    f"… and {len(job.violations) - 50} more violations.",
                    normal))
            elements.append(Spacer(1, 12))

        # Single-line diagram (optional, network export)
        if include_diagrams:
            img_buf = cls._render_diagram(job.network_id)
            if img_buf is not None:
                elements.append(PageBreak())
                elements.append(Paragraph("Single-Line Diagram", h2))
                elements.append(Image(img_buf, width=16 * cm, height=12 * cm))

        # Detailed result tables (truncated)
        for res_name in ("res_bus", "res_line", "res_trafo"):
            rows = (results.get(res_name) or [])[:30]
            if not rows:
                continue
            elements.append(PageBreak())
            elements.append(Paragraph(res_name.replace("_", " ").title(), h2))
            header = list(rows[0].keys())
            data = [header] + [
                [(f"{r[k]:.3f}" if isinstance(r.get(k), float)
                                else str(r.get(k, "")))
                 for k in header]
                for r in rows
            ]
            elements.append(cls._table(data, header_row=True))

        doc.build(elements)

    # -----------------------------------------------------------------
    @staticmethod
    def _table(data, col_widths=None, header_row: bool = False):
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors
        t = Table(data, colWidths=col_widths)
        style = [
            ("FONTNAME",   (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",(0, 0), (-1, -1), 5),
            ("RIGHTPADDING",(0,0), (-1, -1), 5),
        ]
        if header_row:
            style += [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        t.setStyle(TableStyle(style))
        return t

    # -----------------------------------------------------------------
    @staticmethod
    def _render_diagram(network_id: int) -> io.BytesIO | None:
        """Best-effort single-line diagram via matplotlib + pandapower."""
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import pandapower.plotting as plot
            from .pandapower_service import PandapowerService

            net = PandapowerService.build_net_from_db(network_id)
            fig, ax = plt.subplots(figsize=(10, 7))
            plot.simple_plot(net, ax=ax, show_plot=False)
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            return buf
        except Exception:
            return None

    # =================================================================
    #  Excel (openpyxl)
    # =================================================================
    @classmethod
    def _render_xlsx(cls, job: AnalysisJob, out_path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="1F4E79")

        ws["A1"] = "PowerSys Analysis Report"
        ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells("A1:D1")

        meta = [
            ("Job ID",          job.id),
            ("Analysis Type",   job.analysis_type.value),
            ("Status",          job.status.value),
            ("Network ID",      job.network_id),
            ("Converged",       job.converged),
            ("Duration (s)",    job.duration_sec),
            ("Started",         job.started_at.isoformat() if job.started_at else None),
            ("Completed",       job.completed_at.isoformat() if job.completed_at else None),
        ]
        for i, (k, v) in enumerate(meta, start=3):
            ws.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 32

        # Summary block
        summary = (job.results or {}).get("summary") or {}
        if summary:
            start = 3 + len(meta) + 2
            ws.cell(row=start, column=1, value="Summary").font = Font(bold=True, size=14)
            for i, (k, v) in enumerate(summary.items(), start=start + 1):
                ws.cell(row=i, column=1, value=k.replace("_", " ").title()).font = Font(bold=True)
                ws.cell(row=i, column=2, value=v)

        # Sheet 2: Violations
        if job.violations:
            vs = wb.create_sheet("Violations")
            headers = ["Element Type", "PP Index", "Element Name",
                       "Violation Type", "Severity", "Value", "Limit", "Unit", "Message"]
            for c, h in enumerate(headers, start=1):
                cell = vs.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for r, v in enumerate(job.violations, start=2):
                vs.cell(row=r, column=1, value=v.element_type.value)
                vs.cell(row=r, column=2, value=v.element_pp_index)
                vs.cell(row=r, column=3, value=v.element_name)
                vs.cell(row=r, column=4, value=v.violation_type)
                vs.cell(row=r, column=5, value=v.severity.value)
                vs.cell(row=r, column=6, value=v.value)
                vs.cell(row=r, column=7, value=v.limit)
                vs.cell(row=r, column=8, value=v.unit)
                vs.cell(row=r, column=9, value=v.message)
            for col, width in zip("ABCDEFGHI",
                                  [14, 10, 22, 18, 10, 12, 12, 8, 50]):
                vs.column_dimensions[col].width = width

        # Result sheets
        results = job.results or {}
        for res_name in ("res_bus", "res_line", "res_trafo", "res_load",
                         "res_gen", "res_ext_grid", "res_bus_sc",
                         "res_line_sc", "res_trafo_sc"):
            rows = results.get(res_name)
            if not rows:
                continue
            rs = wb.create_sheet(res_name[:31])
            headers = list(rows[0].keys())
            for c, h in enumerate(headers, start=1):
                cell = rs.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, h in enumerate(headers, start=1):
                    rs.cell(row=r_idx, column=c_idx, value=row.get(h))
            for c_idx, _ in enumerate(headers, start=1):
                rs.column_dimensions[
                    rs.cell(row=1, column=c_idx).column_letter
                ].width = 16

        wb.save(out_path)
